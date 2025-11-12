import pandas as pd
import openpyxl  # <-- Usamos openpyxl para bajo consumo de RAM
import sys
from collections import defaultdict

class FinansuenosDataProcessorService:
    """
    Clase responsable de todas las transformaciones de datos.
    OPTIMIZADA (Opción B): Carga las correcciones (84MB)
    usando 'openpyxl' para bajo consumo de RAM.
    """
    def __init__(self, correcciones_path):
        """
        Carga TODAS las hojas de corrección (el archivo de 84MB)
        en la memoria UNA SOLA VEZ, usando 'openpyxl' para
        evitar DataFrames de Pandas gigantes.
        """
        print(f"SERVICE (FINANSUEÑOS): Abriendo (read-only) el archivo de correcciones: {correcciones_path}", flush=True)
        # Cargamos el libro de trabajo en modo 'read_only' (eficiente)
        # 'data_only=True' lee los valores de las fórmulas, no las fórmulas
        try:
            wb = openpyxl.load_workbook(correcciones_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"SERVICE_ERROR: No se pudo abrir el archivo de correcciones '{correcciones_path}' con openpyxl. Error: {e}", flush=True)
            raise e

        # --- Mapa A: Cédulas ---
        print(f"SERVICE (FINANSUEÑOS): 1/5 Creando mapa 'Cedulas a corregir'...", flush=True)
        self.mapa_cedulas = {}
        try:
            ws_cedulas = wb['Cedulas a corregir']
            # Iteramos fila por fila (saltando el header)
            for i, row in enumerate(ws_cedulas.iter_rows(min_row=2)):
                cedula_mal = str(row[0].value).strip()
                cedula_correcta = str(row[1].value).strip()
                if cedula_mal:
                    self.mapa_cedulas[cedula_mal] = cedula_correcta
        except KeyError:
            print("SERVICE_WARN: No se encontró la hoja 'Cedulas a corregir'", flush=True)
        
        # --- Mapa B: Vinculado ---
        print(f"SERVICE (FINANSUEÑOS): 2/5 Creando mapa 'Vinculado'...", flush=True)
        # Necesitamos 4 mapas (NOMBRE, DIRECCI, VINEMAIL, TELEFONO) indexados por CODIGO
        self.mapa_vinc = defaultdict(dict)
        try:
            ws_vinculado = wb['Vinculado']
            # Asumimos: CODIGO(A), NOMBRE(B), DIRECCI(C), VINEMAIL(D), TELEFONO(E)
            # Esto es una suposición, ajusta los índices (ej. row[0]) si es necesario
            # Encuentra los headers
            headers = [cell.value for cell in ws_vinculado[1]]
            idx_codigo = headers.index('CODIGO')
            mapa_cols_vinc = {
                'NOMBRE': headers.index('NOMBRE'),
                'DIRECCI': headers.index('DIRECCI'),
                'VINEMAIL': headers.index('VINEMAIL'),
                'TELEFONO': headers.index('TELEFONO')
            }
            
            for row in ws_vinculado.iter_rows(min_row=2):
                codigo = str(row[idx_codigo].value).strip()
                if codigo:
                    for col_name, col_idx in mapa_cols_vinc.items():
                        self.mapa_vinc[codigo][col_name] = row[col_idx].value
        except Exception as e:
            print(f"SERVICE_WARN: No se encontró o falló la hoja 'Vinculado'. Error: {e}", flush=True)
        
        self.mapa_vinc_cols_df = {'NOMBRE COMPLETO':'NOMBRE', 'DIRECCION DE CORRESPONDENCIA':'DIRECCI', 'CORREO ELECTRONICO':'VINEMAIL', 'CELULAR':'TELEFONO'}

        # --- Mapa C: Tipos de Identificación ---
        print(f"SERVICE (FINANSUEÑOS): 3/5 Creando mapa 'Tipos de identificacion'...", flush=True)
        self.mapa_tipos = {}
        try:
            ws_tipos = wb['Tipos de identificacion']
            # Asumimos: CEDULA CORRECTA(A), CODIGO DATA(B)
            for row in ws_tipos.iter_rows(min_row=2):
                cedula = str(row[0].value).strip()
                codigo_data = row[1].value
                if cedula:
                    self.mapa_tipos[cedula] = codigo_data
        except KeyError:
            print("SERVICE_WARN: No se encontró la hoja 'Tipos de identificacion'", flush=True)

        # --- Mapa D: FNZ001 ---
        print(f"SERVICE (FINANSUEÑOS): 4/5 Creando mapa 'FNZ001'...", flush=True)
        self.mapa_fnz = {}
        try:
            ws_fnz = wb['FNZ001']
            headers_fnz = [cell.value for cell in ws_fnz[1]]
            idx_dsm_tp = headers_fnz.index('DSM_TP')
            idx_dsm_num = headers_fnz.index('DSM_NUM')
            idx_vlr_fnz = headers_fnz.index('VLR_FNZ')
            
            for row in ws_fnz.iter_rows(min_row=2):
                try:
                    vlr_fnz = int(pd.to_numeric(row[idx_vlr_fnz].value, errors='coerce'))
                except:
                    vlr_fnz = 0
                
                if vlr_fnz == 0:
                    continue
                    
                llave_base = str(row[idx_dsm_tp].value).strip() + str(row[idx_dsm_num].value).strip()
                
                # Replicamos la lógica del 'concat'
                facturas = [
                    llave_base.zfill(18),
                    (llave_base + 'C1').zfill(18),
                    (llave_base + 'C2').zfill(18)
                ]
                for fact in facturas:
                    self.mapa_fnz[fact] = vlr_fnz
        except Exception as e:
            print(f"SERVICE_WARN: No se encontró o falló la hoja 'FNZ001'. Error: {e}", flush=True)

        # --- Mapa E: R05 ---
        print(f"SERVICE (FINANSUEÑOS): 5/5 Creando mapa 'R05'...", flush=True)
        self.mapa_r05 = {}
        abonos_sumados = defaultdict(float)
        try:
            ws_r05 = wb['R05']
            headers_r05 = [cell.value for cell in ws_r05[1]]
            idx_tipo = headers_r05.index('MCNTIPCRU2')
            idx_num = headers_r05.index('MCNNUMCRU2')
            idx_abono = headers_r05.index('ABONO')
            
            # Replicamos el 'groupby().sum()'
            for row in ws_r05.iter_rows(min_row=2):
                try:
                    abono = float(pd.to_numeric(row[idx_abono].value, errors='coerce'))
                except:
                    abono = 0.0
                
                if abono == 0.0:
                    continue
                    
                llave_base = str(row[idx_tipo].value).strip() + str(row[idx_num].value).strip()
                abonos_sumados[llave_base] += abono
                
            # Replicamos el 'concat'
            for llave_base, valor_abono in abonos_sumados.items():
                llaves = [
                    llave_base.ljust(20),
                    (llave_base + 'C1').ljust(20),
                    (llave_base + 'C2').ljust(20)
                ]
                for llave in llaves:
                    self.mapa_r05[llave] = valor_abono
        except Exception as e:
            print(f"SERVICE_WARN: No se encontró o falló la hoja 'R05'. Error: {e}", flush=True)

        wb.close() # Cerramos el archivo de 84MB
        print(f"SERVICE (FINANSUEÑOS): Mapas de corrección listos. Archivo Excel cerrado.", flush=True)


    def run_all_transformations(self, chunk_df):
        """
        Ejecuta todos los pasos de limpieza en un 'chunk' (trozo) del DataFrame.
        YA NO CARGA ARCHIVOS. Solo aplica los mapas pre-calculados.
        """
        # print("Servicio: Ejecutando transformaciones en chunk...", flush=True) # Demasiado 'ruido'
        self.df = chunk_df # Asigna el chunk actual a self.df
        
        # Llama a los métodos de transformación
        # (Estos métodos ahora usan los mapas en 'self' en lugar de leer archivos)
        self._correct_data_from_excel() 
        self._update_data_from_sheets()
        self._clean_and_validate_data()
        self._apply_final_formatting()
        
        # print("Servicio: Transformaciones de chunk completadas.", flush=True)
        return self.df

    def _correct_data_from_excel(self):
        """PASO 3: Realiza correcciones (YA NO LEE EXCEL)"""
        # print("  - Corrigiendo desde Excel (memoria)...", flush=True)
        # A. Corregir Cédulas (Usa el mapa)
        self.df['NUMERO DE IDENTIFICACION'] = self.df['NUMERO DE IDENTIFICACION'].replace(self.mapa_cedulas)

        # B. Actualizar campos 'CORREGIR' (Usa el mapa)
        for col_df, col_vinc in self.mapa_vinc_cols_df.items():
            mascara = self.df[col_df].astype(str).str.strip().str.contains('CORREGIR', case=False, na=False)
            if mascara.any():
                ids_a_buscar = self.df.loc[mascara, 'NUMERO DE IDENTIFICACION']
                # Usamos el mapa de diccionarios que creamos
                valores_nuevos = ids_a_buscar.map(lambda x: self.mapa_vinc.get(x, {}).get(col_vinc))
                self.df.loc[mascara, col_df] = valores_nuevos

        # C. Actualizar Tipos de Identificación (Usa el mapa)
        self.df['TIPO DE IDENTIFICACION'] = 1
        self.df['TIPO DE IDENTIFICACION'] = self.df['NUMERO DE IDENTIFICACION'].map(self.mapa_tipos).combine_first(self.df['TIPO DE IDENTIFICACION'])

    def _update_data_from_sheets(self):
        """PASO 4: Actualiza desde FNZ001 y R05 (YA NO LEE EXCEL)"""
        # print("  - Actualizando desde FNZ001 y R05 (memoria)...", flush=True)
        self.df['NUMERO DE LA CUENTA U OBLIGACION'] = self.df['NUMERO DE LA CUENTA U OBLIGACION'].astype(str).str.replace(' ', '').str.zfill(18)
        
        # A. Procesando FNZ001 (Usa el mapa)
        self.df['VALOR INICIAL'] = self.df['NUMERO DE LA CUENTA U OBLIGACION'].map(self.mapa_fnz).combine_first(self.df['VALOR INICIAL'])

        # B. Procesando R05 (Usa el mapa)
        self.df['VALOR SALDO MORA'] = self.df['NUMERO DE LA CUENTA U OBLIGACION'].map(self.mapa_r05).combine_first(self.df['VALOR SALDO MORA'])

    def _clean_and_validate_data(self):
        """PASO 5: Realiza limpieza y validaciones generales."""
        # print("  - Limpiando y validando datos...", flush=True)
        # (Este código no consume mucha RAM, se queda igual)
        # A. Limpieza de caracteres de texto
        letter_replacements = {'Ñ':'N','Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','Ü':'U','Ÿ':'Y','Â':'A','Ã':'A','š':'S','©':'C','ñ':'N','á':'A','é':'E','í':'I','ó':'O','ú':'U','ü':'U','ÿ':'Y','â':'A','ã':'A'}
        chars_to_remove = ['@','°','|','¬','¡','“','#','$','%','&','/','(',')','=','‘','\\','¿','+','~','´´','´','[','{','^','-','_','.',':',',',';','<','>','Æ','±']
        string_cols = self.df.select_dtypes(include='object').columns.drop('CORREO ELECTRONICO', errors='ignore')
        for col in string_cols:
            self.df[col] = self.df[col].astype(str)
            for old, new in letter_replacements.items(): self.df[col] = self.df[col].str.replace(old, new, regex=False)
            for char in chars_to_remove: self.df[col] = self.df[col].str.replace(char, '', regex=False)
        
        # B. Limpieza y validación de fechas
        for col in ["FECHA APERTURA", "FECHA VENCIMIENTO", "FECHA DE PAGO"]:
            self.df[col] = pd.to_numeric(self.df[col], errors='coerce').fillna(0).astype('Int64').astype(str)
        condicion_fecha_invalida = self.df['FECHA VENCIMIENTO'] < self.df['FECHA APERTURA']
        self.df.loc[condicion_fecha_invalida, 'FECHA VENCIMIENTO'] = self.df['FECHA APERTURA']
        mascara_fecha_pago = (self.df['FECHA DE PAGO'].str.upper().str.contains('NA')) | (self.df['FECHA DE PAGO'] == '0')
        self.df.loc[mascara_fecha_pago, 'FECHA DE PAGO'] = '00000000'

        # C. Limpieza y validación de valores numéricos
        columnas_numericas = ["VALOR INICIAL", "VALOR SALDO DEUDA", "VALOR DISPONIBLE", "V CUOTA MENSUAL", "VALOR SALDO MORA"]
        for col in columnas_numericas:
            self.df[col] = pd.to_numeric(self.df[col], errors='coerce').fillna(0)
            if col != 'VALOR DISPONIBLE':
                self.df.loc[self.df[col] < 10000, col] = 0
        self.df['VALOR DISPONIBLE'] = 0
        self.df[columnas_numericas] = self.df[columnas_numericas].astype(int)


    def _apply_final_formatting(self):
        """PASO 6: Aplica el formato final de texto y longitud."""
        # print("  - Aplicando formatos finales...", flush=True)
        # (Este código no consume mucha RAM, se queda igual)
        self.df['NOMBRE COMPLETO'] = self.df['NOMBRE COMPLETO'].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip().str.upper()
        replacements_map = {'1118291452':'FANDINO LAYNE ASTRID', '1025529458':'MARTINEZ MUNOZ JOSE MANUEL', '25559122':'RAMIREZ DE CASTRO MARIA ESTELLA'}
        for id_number, new_name in replacements_map.items():
            self.df.loc[self.df['NUMERO DE IDENTIFICACION'] == id_number, 'NOMBRE COMPLETO'] = new_name
        
        col_ciudad = 'CIUDAD CORRESPONDENCIA'
        self.df[col_ciudad] = self.df[col_ciudad].astype(str).fillna('')
        mascara_reemplazo_ciudad = (self.df[col_ciudad].str.strip() == '') | (self.df[col_ciudad].str.strip().str.upper() == 'N/A') | (self.df[col_ciudad].str.strip() == '0')
        self.df.loc[mascara_reemplazo_ciudad, col_ciudad] = 'POPAYAN'

        col_celular = 'CELULAR'
        self.df[col_celular] = self.df[col_celular].astype(str).str.replace(r'\D', '', regex=True)
        es_fijo_valido = (self.df[col_celular].str.len() == 7)
        es_celular_valido = (self.df[col_celular].str.len() == 10) & (self.df[col_celular].str.startswith('3'))
        self.df.loc[~(es_fijo_valido | es_celular_valido), col_celular] = ''
        
        col_email = 'CORREO ELECTRONICO'
        self.df[col_email] = self.df[col_email].astype(str).str.strip()
        placeholders_a_eliminar = ['CORREGIR', 'PENDIENTE', 'NOTIENE', 'SINC', 'NN@', 'AAA@']
        for placeholder in placeholders_a_eliminar:
            self.df.loc[self.df[col_email].str.contains(placeholder, case=False, na=False), col_email] = ''
        email_regex_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        self.df.loc[~self.df[col_email].str.match(email_regex_pattern, na=False), col_email] = ''

        # Formatos de longitud y valores fijos
        self.df['ESTADO ORIGEN DE LA CUENTA'] = '0'
        self.df['RESPONSABLE'] = self.df['RESPONSABLE'].astype(str).str.zfill(2)
        self.df['NOVEDAD'] = self.df['NOVEDAD'].astype(str).str.zfill(2)
        self.df['TOTAL CUOTAS'] = self.df['TOTAL CUOTAS'].astype(str).str.zfill(3)
        self.df['CUOTAS CANCELADAS'] = self.df['CUOTAS CANCELADAS'].astype(str).str.zfill(3)
        self.df['CUOTAS EN MORA'] = self.df['CUOTAS EN MORA'].astype(str).str.zfill(3)
        self.df['FECHA LIMITE DE PAGO'] = self.df['FECHA LIMITE DE PAGO'].astype(str)
        self.df['SITUACION DEL TITULAR'] = '0'
        self.df['EDAD DE MORA'] = self.df['EDAD DE MORA'].astype(str).str.zfill(3)
        self.df['FORMA DE PAGO'] = self.df['FORMA DE PAGO'].astype(str)
        self.df['FECHA ESTADO ORIGEN'] = self.df['FECHA ESTADO ORIGEN'].astype(str)
        self.df['ESTADO DE LA CUENTA'] = self.df['ESTADO DE LA CUENTA'].astype(str).str.zfill(2)
        self.df['FECHA ESTADO DE LA CUENTA'] = self.df['FECHA ESTADO DE LA CUENTA'].astype(str)
        self.df['ADJETIVO'] = '0'
        self.df['FECHA DE ADJETIVO'] = self.df['FECHA DE ADJETIVO'].astype(str).str.zfill(8)
        self.df['CLAUSULA DE PERMANENCIA'] = self.df['CLAUSULA DE PERMANENCIA'].astype(str).str.zfill(3)
        self.df['FECHA CLAUSULA DE PERMANENCIA'] = self.df['FECHA CLAUSULA DE PERMANENCIA'].astype(str).str.zfill(8)
        
        self.df['NOMBRE COMPLETO'] = self.df['NOMBRE COMPLETO'].str.ljust(45)
        self.df['DIRECCION DE CORRESPONDENCIA'] = self.df['DIRECCION DE CORRESPONDENCIA'].astype(str).str.ljust(60)
        self.df['CIUDAD CORRESPONDENCIA'] = self.df[col_ciudad].str.ljust(20)
        self.df['CORREO ELECTRONICO'] = self.df[col_email].str.ljust(60)
        self.df['CELULAR'] = self.df[col_celular].str.zfill(12)
        self.df['NUMERO DE IDENTIFICACION'] = self.df['NUMERO DE IDENTIFICACION'].str.zfill(11)
        self.df['TIPO DE IDENTIFICACION'] = self.df['TIPO DE IDENTIFICACION'].astype(int).astype(str)