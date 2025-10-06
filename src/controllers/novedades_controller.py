import datetime
from tkinter import filedialog, messagebox
import numpy as np
import pandas as pd
from pathlib import Path
from src.services.novedades.novedades_service import NovedadesService
from src.services.novedades.analisis_service import AnalisisService
from src.services.novedades.recaudo_service import RecaudoR91Service 
from src.services.novedades.franjas_service import ReporteFranjasService
from src.models.novedad_model import configuracion

class NovedadesAnalisisController:
    def __init__(self):
        self.view = None
        # La ruta al caché se define aquí para que el controlador la conozca
        self.cache_path = Path(__file__).resolve().parent.parent.parent / "cache" / "reporte_base_mensual.feather"

    def set_view(self, view):
        """Asigna la vista a este controlador."""
        self.view = view

    def _cargar_reporte_base(self, ruta_base, config):
        """
        Carga el archivo de reporte base, permitiendo que pandas detecte
        automáticamente los tipos de datos de fecha y texto.
        """
        print("🔄 Cargando reporte base con configuración del modelo...")
        
        config_base = config.get("BASE_MENSUAL", {})
        mapa_de_tipos = config_base.get("dtype_map", None)

        if not mapa_de_tipos:
            messagebox.showwarning("Configuración Faltante", "No se encontró el 'dtype_map' en la configuración.")
            return pd.read_excel(ruta_base)
        
        df_base = pd.read_excel(ruta_base, dtype=mapa_de_tipos)
        
        print("✅ Reporte base cargado exitosamente con tipos de datos correctos.")
        return df_base

    def _cargar_y_unir_archivos(self, file_paths, config_key):
        """
        Función interna para leer y concatenar múltiples archivos,
        seleccionando el motor de Excel correcto.
        """
        if not file_paths:
            return pd.DataFrame()
        
        df_list = []
        file_config = configuracion[config_key]
        for path in file_paths:
            # --- LÓGICA CORREGIDA PARA SELECCIONAR EL MOTOR ---
            engine_to_use = None
            # Convertimos la ruta a minúsculas para una comparación segura
            file_ext = path.lower() 

            if file_ext.endswith('.xlsx'):
                engine_to_use = 'openpyxl'
            elif file_ext.endswith('.xls'):
                engine_to_use = 'xlrd'
            
            # Le pasamos el motor correcto a pandas al leer el archivo
            df = pd.read_excel(
                path, 
                engine=engine_to_use, # <-- Se añade el motor
                usecols=file_config["usecols"]
            ).rename(columns=file_config["rename_map"])

            if config_key == "NOVEDADES":
                print(f"🔍 Identificando empresa para el archivo: {Path(path).name}")
                if 'arpesod' in file_ext:
                    df['Empresa'] = 'ARPESOD'
                elif 'finansueños' in file_ext:
                    df['Empresa'] = 'FINANSUEÑOS'
                else:
                    df['Empresa'] = 'OTRA'

            df_list.append(df)
        
        return pd.concat(df_list, ignore_index=True)

    def _escribir_y_formatear_franjas(self, writer, df, sheet_name):
        """
        Función definitiva que escribe y formatea la hoja de franjas manualmente
        para evitar el bug de pandas/openpyxl.
        """
        print(f"✍️ Escribiendo y formateando la hoja '{sheet_name}' manualmente...")
        
        workbook = writer.book
        worksheet = workbook.create_sheet(sheet_name)
        
        from openpyxl.styles import Alignment, Font
        alineacion_centrada = Alignment(horizontal='center', vertical='center')
        fuente_negrita = Font(bold=True)

        # 1. Escribir y formatear los encabezados (filas 1 y 2)
        header_groups = {}
        for col_idx, col_tuple in enumerate(df.columns, 1):
            top_header = col_tuple[0]
            sub_header = col_tuple[1]
            
            # Escribir los valores en las celdas
            cell1 = worksheet.cell(row=1, column=col_idx, value=top_header)
            cell2 = worksheet.cell(row=2, column=col_idx, value=sub_header)
            
            # Aplicar estilo a ambas celdas del encabezado
            cell1.font = fuente_negrita
            cell1.alignment = alineacion_centrada
            cell2.font = fuente_negrita
            cell2.alignment = alineacion_centrada
            
            # Guardar el rango de cada encabezado superior para unirlo después
            if top_header not in header_groups:
                header_groups[top_header] = {'start': col_idx, 'end': col_idx}
            else:
                header_groups[top_header]['end'] = col_idx
        
        # 2. Aplicar merges a los encabezados
        # Merges horizontales (para las franjas)
        for group_info in header_groups.values():
            if group_info['start'] < group_info['end']:
                worksheet.merge_cells(start_row=1, end_row=1, start_column=group_info['start'], end_column=group_info['end'])
        
        # Merges verticales (para ZONA, REGIONAL, etc.)
        for col_idx, col_tuple in enumerate(df.columns, 1):
            if col_tuple[1] == '':
                worksheet.merge_cells(start_row=1, end_row=2, start_column=col_idx, end_column=col_idx)

        # 3. Escribir los datos, fila por fila (empezando en la fila 3)
        for row_idx, data_row in enumerate(df.itertuples(index=False), 3):
            for col_idx, value in enumerate(data_row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                # Si el valor es un porcentaje en formato texto, se alinea al centro
                if isinstance(value, str) and '%' in value:
                    cell.alignment = alineacion_centrada
        
        # 4. Aplicar merges a las celdas de datos por Zona
        cols_to_merge = [1, 15, 16] # ZONA, Total_Recaudo, Recaudo_Anticipo
        start_row = 3
        
        if worksheet.max_row >= start_row:
            current_zone_value = worksheet.cell(row=start_row, column=1).value
            for row_idx in range(start_row + 1, worksheet.max_row + 2):
                next_zone_value = worksheet.cell(row=row_idx, column=1).value
                if next_zone_value != current_zone_value:
                    end_row = row_idx - 1
                    if start_row < end_row:
                        for col in cols_to_merge:
                            worksheet.merge_cells(start_row=start_row, end_row=end_row, start_column=col, end_column=col)
                            worksheet.cell(row=start_row, column=col).alignment = alineacion_centrada
                    start_row = row_idx
                    current_zone_value = next_zone_value

    def procesar_archivos(self, rutas_novedades, ruta_base, rutas_analisis, rutas_r91, ruta_usuarios):
        """
        Orquesta todo el proceso: carga el caché, aplica novedades, calcula el rodamiento
        y guarda un reporte multi-hoja.
        """
        if not rutas_novedades or not rutas_analisis or not rutas_r91:
            messagebox.showwarning("Archivos Faltantes", "Debes seleccionar los archivos de Novedades y Análisis.")
            return
            
        try:
            # --- CAMBIO: Cargar la base desde el archivo Excel seleccionado ---
            print(f"🔄 Cargando reporte base desde: {ruta_base}")
            # Usamos dtype=str para evitar problemas de formato de Excel con las cédulas
            df_base = self._cargar_reporte_base(ruta_base, configuracion)
            
            # Si la carga falla por alguna razón, el método nuevo podría retornar None
            if df_base is None or df_base.empty:
                messagebox.showerror("Error", "No se pudo cargar el reporte base correctamente.")
                return

            # 1. Cargar y unir todos los archivos de entrada
            df_novedades_unido = self._cargar_y_unir_archivos(rutas_novedades, "NOVEDADES")
            df_analisis_unido = self._cargar_y_unir_archivos(rutas_analisis, "ANALISIS")
            df_r91_unido = self._cargar_y_unir_archivos(rutas_r91, "R91")
            df_usuarios_unido = self._cargar_y_unir_archivos(ruta_usuarios, "USUARIOS")            
            # --- INICIO DE LA SOLUCIÓN: Limpieza de Llaves ---
            print("🧹 Estandarizando y limpiando llaves de unión en todos los archivos...")

            # Limpiar Cedula_Cliente en df_base y df_novedades
            if 'Cedula_Cliente' in df_base.columns:
                df_base['Cedula_Cliente'] = df_base['Cedula_Cliente'].astype(str).str.strip()
            if 'Cedula_Cliente' in df_novedades_unido.columns:
                df_novedades_unido['Cedula_Cliente'] = df_novedades_unido['Cedula_Cliente'].astype(str).str.strip()
            
            # Crear y limpiar la llave 'Credito' en los archivos que la usan
            if not df_analisis_unido.empty:
                df_analisis_unido['Credito'] = df_analisis_unido['Tipo_Credito'].astype(str) + '-' + df_analisis_unido['Numero_Credito'].astype(str)
                df_analisis_unido['Credito'] = df_analisis_unido['Credito'].str.strip()

            if not df_r91_unido.empty:
                df_r91_unido['Credito'] = df_r91_unido['Tipo_Credito'].astype(str) + '-' + df_r91_unido['Numero_Credito'].astype(str)
                df_r91_unido['Credito'] = df_r91_unido['Credito'].str.strip()
            
            # Limpiar la llave 'Usuario_Novedad' (esto ya lo tenías, pero es parte de la misma lógica)
            if 'Usuario_Novedad' in df_novedades_unido.columns:
                df_novedades_unido['Usuario_Novedad'] = df_novedades_unido['Usuario_Novedad'].astype(str).str.strip().str.lower()
            if not df_usuarios_unido.empty:
                df_usuarios_unido['Usuario_Novedad'] = df_usuarios_unido['Usuario_Novedad'].astype(str).str.strip().str.lower()
            # --- FIN DE LA SOLUCIÓN ---
            
            # 2. Aplicar Novedades (Ahora usará las cédulas limpias)
            novedades_service = NovedadesService(configuracion)
            df_base_enriquecido, df_novedades_detallado = novedades_service.aplicar_novedades(df_base, df_novedades_unido)
            
            print("🔧 Asegurando compatibilidad de tipos para la unión de usuarios...") 
            if not df_novedades_detallado.empty:
                df_novedades_detallado['Usuario_Novedad'] = df_novedades_detallado['Usuario_Novedad'].astype(str).str.strip().str.lower()
            
            if not df_usuarios_unido.empty:
                df_usuarios_unido['Usuario_Novedad'] = df_usuarios_unido['Usuario_Novedad'].astype(str).str.strip().str.lower()
            
            df_novedades_detallado = pd.merge(
                df_novedades_detallado, 
                df_usuarios_unido, 
                on='Usuario_Novedad', 
                how='left'
            )
            
            columnas_a_mayusculas = ['Usuario_Novedad', 'Nombre_Usuario', 'Cargo_Usuario']
            for col in columnas_a_mayusculas:
                # Verificamos si la columna existe en el DataFrame para evitar errores
                if col in df_novedades_detallado.columns:
                    # Usamos .astype(str) por si hay algún valor no textual (como NaN) y luego .str.upper()
                    df_novedades_detallado[col] = df_novedades_detallado[col].astype(str).str.upper()
                    
            # 3. Calcular Rodamiento
            analisis_service = AnalisisService(configuracion)
            df_con_rodamiento = analisis_service.calcular_rodamiento(df_base_enriquecido, df_analisis_unido)

            # 4. Calcular Recaudos
            recaudo_service = RecaudoR91Service()
            df_recaudos = recaudo_service.procesar_recaudos(df_r91_unido)

            # 5. Unir la información de recaudos al reporte final
            df_final = pd.merge(df_con_rodamiento, df_recaudos, on='Credito', how='left')
            for col in ['Recaudo_Anticipado', 'Recaudo_Meta', 'Total_Recaudo']:
                if col in df_final.columns:
                    df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna(0)
            
            # 1. Formatear las columnas de FECHAS PURAS para quitar la hora.
            columnas_fecha_puras_sin_hora = ['Fecha_Desembolso', 'Fecha_Facturada','Fecha_Ultimo_pago']
            for col in columnas_fecha_puras_sin_hora:
                if col in df_final.columns:
                    # La parte .dt.date elimina la hora
                    df_final[col] = pd.to_datetime(df_final[col], errors='coerce').dt.date

            # 2. Formatear las columnas de FECHAS MIXTAS (FECHA y TEXTO) para quitar la hora.
            columnas_fecha_mixtas = ['Fecha_Cuota_Vigente', 'Fecha_Cuota_Atraso']
            for col in columnas_fecha_mixtas:
                if col in df_final.columns:
                    # La parte .date() elimina la hora solo de los objetos de fecha
                    df_final[col] = df_final[col].apply(
                        lambda x: x.date() if isinstance(x, (pd.Timestamp, datetime.datetime)) else x
                    )
            
             # PASO DE VERIFICACIÓN (OPCIONAL): Puedes descomentar esto para depurar
            print("Tipos de datos en 'Fecha_Cuota_Vigente' antes de guardar:")
            print(df_final['Fecha_Cuota_Vigente'].apply(type).value_counts())
            # --- FIN DEL BLOQUE CORREGIDO ---
            
            orden_columnas_analisis = [
                'Empresa', 'Credito', 'Fecha_Desembolso', 'Factura_Venta', 'Fecha_Facturada',
                'Nombre_Producto', 'Cantidad_Producto', 'Obsequio', 'Cantidad_Obsequio',
                'Cantidad_Total_Producto', 'Cedula_Cliente', 'Nombre_Cliente', 'Correo',
                'Celular', 'Direccion', 'Barrio', 'Nombre_Ciudad', 'Zona', 'Cobrador',
                'Telefono_Cobrador', 'Zona_Cobro', 'Call_Center_Apoyo', 'Nombre_Call_Center',
                'Telefono_Call_Center', 'Regional_Cobro', 'Gestor', 'Telefono_Gestor',
                'Jefe_ventas', 'Codigo_Vendedor', 'Nombre_Vendedor', 'Movil_Vendedor',
                'Vendedor_Activo', 'Lider_Zona', 'Movil_Lider', 'Codigo_Centro_Costos',
                'Regional_Venta', 'Codeudor1', 'Nombre_Codeudor1', 'Telefono_Codeudor1',
                'Ciudad_Codeudor1', 'Codeudor2', 'Nombre_Codeudor2', 'Telefono_Codeudor2',
                'Ciudad_Codeudor2', 'Valor_Desembolso', 'Total_Cuotas', 'Valor_Cuota',
                'Dias_Atraso', 'Franja_Meta','Franja_Cartera', 'Saldo_Capital', 'Saldo_Interes_Corriente',
                'Saldo_Avales', 'Meta_Intereses', 'Meta_General','Meta_Saldo', 'Meta_%', 'Meta_$',
                'Meta_T.R_%', 'Meta_T.R_$', 'Cuotas_Pagadas', 'Cuota_Vigente',
                'Fecha_Cuota_Vigente', 'Valor_Cuota_Vigente', 'Fecha_Cuota_Atraso',
                'Primera_Cuota_Mora', 'Valor_Cuota_Atraso', 'Valor_Vencido',
                'Fecha_Ultima_Novedad', 'Cantidad_Novedades','Fecha_Ultimo_pago','Rango_Ultimo_pago', 'Dias_Atraso_Final',
                'Franja_Meta_Final','Franja_Cartera_Final', 'Rodamiento','Rodamiento_Cartera' ,
                'Recaudo_Anticipado', 'Recaudo_Meta','Total_Recaudo','Total_Recaudo_Sin_Anti'
            ]

            orden_columnas_detalle = [
                'Empresa','Cedula_Cliente', 'Nombre_Cliente', 'Fecha_Novedad', 'Usuario_Novedad',
                'Nombre_Usuario', 'Cargo_Usuario', 'Celular_Corporativo','Codigo_Novedad', 'Tipo_Novedad',
                'Novedad', 'Fecha_Compromiso', 'Valor'
            ]

            df_final = df_final[[col for col in orden_columnas_analisis if col in df_final.columns]]
            df_novedades_detallado = df_novedades_detallado[[col for col in orden_columnas_detalle if col in df_novedades_detallado.columns]]
            
            franjas_service = ReporteFranjasService()
            df_reporte_franjas = franjas_service.generar_reporte(df_final)

            # 6. Guardar el reporte multi-hoja
            ruta_salida = filedialog.asksaveasfilename(
                defaultextension=".xlsx", 
                initialfile="Reporte_Final_Analisis.xlsx"
            )
            if not ruta_salida: return

            print("💾 Guardando datos en el archivo Excel...")
            with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
                # Guardamos las hojas simples de forma normal
                df_final.to_excel(writer, sheet_name='Analisis_de_Cartera', index=False)
                df_novedades_detallado.to_excel(writer, sheet_name='Detalle_Novedades', index=False)
                
                # Usamos nuestra nueva función manual para la hoja compleja
                self._escribir_y_formatear_franjas(writer, df_reporte_franjas, 'Reporte_Franjas')
            
            print("✅ Reporte final con formato guardado exitosamente.")
            messagebox.showinfo("Éxito", f"Reporte unificado guardado exitosamente en:\n{ruta_salida}")

        except Exception as e:
            messagebox.showerror("Error en el Proceso", f"Ocurrió un error general:\n{e}")